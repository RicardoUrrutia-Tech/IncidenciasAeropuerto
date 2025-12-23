import streamlit as st
import pandas as pd
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

st.set_page_config(page_title="Incidencias / Ausentismo / Asistencia", layout="wide")
st.title("App Incidencias / Ausentismo / Asistencia")

# =========================
# Constantes / Estilo
# =========================
CABIFY = {
    "m1": "1F123F",
    "m2": "362065",
    "m3": "4A2B8D",
    "m4": "5B34AC",
    "m5": "7145D6",
    "m6": "8A6EE4",
    "m7": "A697ED",
    "m8": "C4BDF5",
    "m9": "DFDAF8",
    "m10": "F5F1FC",
    "m11": "FAF8FE",
    "pink": "E83C96",
    "red": "E74A41",
    "orange": "EA8C2E",
    "yellow": "EFBD03",
    "blue": "4583D4",
    "green": "0C936B",  # ojo: era "OC936B" pero debe ser 0 (cero)
}

CLASIF_OPTS = ["Seleccionar", "Injustificada", "Permiso", "No Procede - Cambio de Turno"]

# =========================
# Helpers
# =========================
def normalize_rut(x) -> str:
    if pd.isna(x):
        return ""
    return str(x).strip().upper().replace(".", "").replace(" ", "")

def try_parse_date_any(x):
    if pd.isna(x):
        return pd.NaT
    return pd.to_datetime(x, errors="coerce", dayfirst=True)

def excel_to_df(file, sheet_index=0):
    return pd.read_excel(file, sheet_name=sheet_index, engine="openpyxl")

def find_col(df: pd.DataFrame, candidates):
    norm_map = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        key = str(cand).strip().lower()
        if key in norm_map:
            return norm_map[key]
    return None

def get_num(df, candidates):
    col = find_col(df, candidates if isinstance(candidates, list) else [candidates])
    if not col:
        return pd.Series([0.0] * len(df))
    return pd.to_numeric(df[col], errors="coerce").fillna(0.0)

def safe_text_series(df, candidates, default=""):
    col = find_col(df, candidates)
    if not col:
        return pd.Series([default] * len(df))
    return df[col].astype(str).fillna(default)

def maybe_filter_area(df, only_area: str):
    if not only_area:
        return df
    area_col = find_col(df, ["Área", "Area", "AREA"])
    if not area_col:
        return df
    return df[df[area_col].astype(str).str.upper().str.contains(only_area.upper(), na=False)].copy()

# =========================
# Excel builder (dinámico)
# =========================
def style_ws_cabify(ws):
    header_fill = PatternFill("solid", fgColor=CABIFY["m4"])
    header_font = Font(color="FFFFFF", bold=True)
    body_fill = PatternFill("solid", fgColor=CABIFY["m11"])
    alt_fill = PatternFill("solid", fgColor=CABIFY["m10"])

    thin = Side(style="thin", color=CABIFY["m7"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.fill = alt_fill if (row_idx % 2 == 0) else body_fill

    # ancho automático simple
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 55)

    ws.freeze_panes = "A2"

def write_df_to_sheet(wb, sheet_name, df: pd.DataFrame):
    ws = wb.create_sheet(title=sheet_name[:31])
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    return ws

def set_date_format(ws, df_cols, col_name="Fecha"):
    if col_name not in df_cols:
        return
    idx = list(df_cols).index(col_name) + 1
    for r in range(2, ws.max_row + 1):
        c = ws.cell(row=r, column=idx)
        c.number_format = "dd-mm-yyyy"

def add_list_sheet(wb):
    ws = wb.create_sheet(title="Listas")
    ws["A1"] = "Clasificación Manual"
    for i, opt in enumerate(CLASIF_OPTS, start=2):
        ws[f"A{i}"] = opt
    style_ws_cabify(ws)
    ws.column_dimensions["A"].width = 35
    return ws

def apply_dropdown_range(ws_target, col_letter, start_row, end_row, list_ws_name="Listas", list_range="A2:A5"):
    # Data Validation con referencia a rango (mejor para Google Sheets)
    dv = DataValidation(
        type="list",
        formula1=f"={list_ws_name}!${list_range.split(':')[0]}:${list_range.split(':')[1]}",
        allow_blank=False
    )
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")
    ws_target.add_data_validation(dv)

def build_resumen_sheet(wb, incidencias_ws_name="Incidencias"):
    ws = wb.create_sheet(title="Resumen")
    headers = ["Tipo_Incidencia", "Cantidad Injustificada", "Cantidad Permiso", "Cantidad No Procede - Cambio de Turno", "Total"]
    ws.append(headers)

    tipos = ["Marcaje/Turno", "Inasistencia"]
    # Asumimos columnas en Incidencias: Tipo_Incidencia (I), Clasificación Manual (K)
    # Pero para no depender de letras fijas, definimos aquí en base a nuestro dataframe exportado:
    # Fecha(A), Nombre(B), Primer Apellido(C), Segundo Apellido(D), RUT(E), Turno(F),
    # Especialidad(G), Supervisor(H), Tipo_Incidencia(I), Detalle(J), Clasificación Manual(K)

    for t in tipos:
        row = [t]
        # COUNTIFS(Incidencias!$I:$I,"Marcaje/Turno",Incidencias!$K:$K,"Injustificada")
        row.append(f'=COUNTIFS({incidencias_ws_name}!$I:$I,"{t}",{incidencias_ws_name}!$K:$K,"Injustificada")')
        row.append(f'=COUNTIFS({incidencias_ws_name}!$I:$I,"{t}",{incidencias_ws_name}!$K:$K,"Permiso")')
        row.append(f'=COUNTIFS({incidencias_ws_name}!$I:$I,"{t}",{incidencias_ws_name}!$K:$K,"No Procede - Cambio de Turno")')
        row.append("=SUM(B{r}:D{r})".format(r=ws.max_row + 1))
        ws.append(row)

    # Total general
    ws.append(["TOTAL",
               f"=SUM(B2:B{1+len(tipos)})",
               f"=SUM(C2:C{1+len(tipos)})",
               f"=SUM(D2:D{1+len(tipos)})",
               f"=SUM(E2:E{1+len(tipos)})"])
    style_ws_cabify(ws)
    return ws

def build_cumplimiento_sheet(wb, turnos_df: pd.DataFrame, incidencias_ws_name="Incidencias"):
    """
    turnos_df debe traer:
    - RUT_norm (sin puntos)
    - Turnos_planificados
    - Nombre, Primer Apellido, Segundo Apellido (opcional)
    Dejamos fórmulas que cuentan injustificadas por RUT en hoja Incidencias y calculan %.
    """
    ws = wb.create_sheet(title="Cumplimiento")

    headers = [
        "Nombre", "Primer Apellido", "Segundo Apellido",
        "RUT_norm", "Turnos_planificados",
        "Injustificadas",
        "Cumplimiento_%"
    ]
    ws.append(headers)

    # Incidencias: RUT está en E, Clasificación Manual en K
    for _, r in turnos_df.iterrows():
        nombre = r.get("Nombre", "")
        pa = r.get("Primer Apellido", "")
        sa = r.get("Segundo Apellido", "")
        rut_norm = r.get("RUT_norm", "")
        turnos = int(r.get("Turnos_planificados", 0)) if pd.notna(r.get("Turnos_planificados", 0)) else 0

        next_row = ws.max_row + 1
        ws.append([nombre, pa, sa, rut_norm, turnos, "", ""])

        # Injustificadas: COUNTIFS(Incidencias!$E:$E, rut_norm_con_puntos? -> en Incidencias guardamos RUT original con puntos)
        # Para hacerlo robusto, en Incidencias agregamos también una columna "RUT_norm" NO (no la quieres visible).
        # Como NO la quieres en export, aquí contamos por coincidencia parcial en texto: sustituimos puntos/guión sería ideal, pero Excel no es lindo.
        # Solución práctica: contamos por coincidencia "contiene" no es posible con COUNTIFS.
        # Por eso en Incidencias exportaremos RUT "normalizado" en una columna oculta NO; entonces en Excel podemos contar exacto.
        # (Esa columna la ocultamos en el Excel final, no aparece para el usuario, pero habilita fórmulas.)
        # => En el export pondremos columna "RUT_norm_hidden" al final y la ocultamos.
        ws.cell(row=next_row, column=6).value = f'=COUNTIFS({incidencias_ws_name}!$L:$L, "{rut_norm}", {incidencias_ws_name}!$K:$K, "Injustificada")'
        # Cumplimiento_% = MAX(0, (1 - Injustificadas/Turnos_planificados) * 100)
        ws.cell(row=next_row, column=7).value = f'=IFERROR(MAX(0,(1 - F{next_row}/E{next_row})*100),100)'

    style_ws_cabify(ws)
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 16
    return ws

def to_excel_bytes_dynamic(incidencias_df: pd.DataFrame, turnos_base_df: pd.DataFrame):
    """
    Crea Excel con:
    - Listas (dropdown)
    - Incidencias (editable + dropdown + fecha)
    - Resumen (fórmulas)
    - Cumplimiento (fórmulas)
    """
    output = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    ws_list = add_list_sheet(wb)

    # Incidencias
    ws_inc = write_df_to_sheet(wb, "Incidencias", incidencias_df)
    set_date_format(ws_inc, incidencias_df.columns, "Fecha")

    # Dropdown en Clasificación Manual (col K)
    if "Clasificación Manual" in incidencias_df.columns:
        idx = list(incidencias_df.columns).index("Clasificación Manual") + 1
        col_letter = ws_inc.cell(row=1, column=idx).column_letter
        end_row = max(2, ws_inc.max_row)
        apply_dropdown_range(ws_inc, col_letter, 2, end_row, list_ws_name="Listas", list_range="A2:A5")

    style_ws_cabify(ws_inc)

    # Resumen (fórmulas)
    ws_res = build_resumen_sheet(wb, incidencias_ws_name="Incidencias")

    # Cumplimiento (fórmulas)
    ws_cump = build_cumplimiento_sheet(wb, turnos_base_df, incidencias_ws_name="Incidencias")

    # Ocultar hoja Listas (pero queda para validación)
    ws_list.sheet_state = "hidden"

    # Ocultar la columna RUT_norm_hidden en Incidencias si existe (col L)
    if "RUT_norm_hidden" in incidencias_df.columns:
        idx = list(incidencias_df.columns).index("RUT_norm_hidden") + 1
        col_letter = ws_inc.cell(row=1, column=idx).column_letter
        ws_inc.column_dimensions[col_letter].hidden = True

    wb.save(output)
    output.seek(0)
    return output

# =========================
# UI Inputs
# =========================
with st.sidebar:
    st.header("Cargar archivos (Excel)")
    f_turnos = st.file_uploader("1) Codificación Turnos BUK", type=["xlsx"])
    f_reporte_turnos = st.file_uploader("2) Reporte Turnos (Activos + Turnos)", type=["xlsx"])
    f_detalle = st.file_uploader("3) Detalle Turnos Colaboradores (Hoja1=Inasistencias, Hoja2=Asistencias)", type=["xlsx"])

    st.divider()
    st.subheader("Filtros")
    only_area = st.text_input("Filtrar Área (opcional)", value="AEROPUERTO")
    min_horas = st.number_input(
        "Tiempo mínimo de incidencia (horas)",
        min_value=0.0,
        value=0.0,
        step=0.25,
        help="Se aplica a asistencias: (Retraso + Salida Anticipada) >= este umbral."
    )

if not all([f_turnos, f_reporte_turnos, f_detalle]):
    st.info("Sube los 3 archivos para comenzar.")
    st.stop()

# =========================
# Load
# =========================
df_turnos = excel_to_df(f_turnos, 0)  # queda para reglas futuras (mapeo sigla-horario, etc.)
df_activos = excel_to_df(f_reporte_turnos, 0)

# Detalle Turnos Colaboradores:
df_inasist = excel_to_df(f_detalle, 0)  # Hoja 1
df_asist = excel_to_df(f_detalle, 1)    # Hoja 2

# =========================
# Normalización columnas
# =========================
rut_col_inas = find_col(df_inasist, ["RUT", "Rut", "rut"])
rut_col_as = find_col(df_asist, ["RUT", "Rut", "rut"])

if not rut_col_inas or not rut_col_as:
    st.error("No pude detectar la columna RUT en una de las hojas del 'Detalle Turnos Colaboradores'.")
    st.stop()

df_inasist["RUT_norm"] = df_inasist[rut_col_inas].apply(normalize_rut)
df_asist["RUT_norm"] = df_asist[rut_col_as].apply(normalize_rut)

# Fechas base
dia_col_inas = find_col(df_inasist, ["Día", "Dia", "DIA", "día"])
df_inasist["Fecha_base"] = df_inasist[dia_col_inas].apply(try_parse_date_any) if dia_col_inas else pd.NaT

fecha_ent_col_as = find_col(df_asist, ["Fecha Entrada", "Fecha_Entrada", "Fecha entrada"])
dia_col_as = find_col(df_asist, ["Día", "Dia", "DIA", "día"])
if fecha_ent_col_as:
    df_asist["Fecha_base"] = df_asist[fecha_ent_col_as].apply(try_parse_date_any)
elif dia_col_as:
    df_asist["Fecha_base"] = df_asist[dia_col_as].apply(try_parse_date_any)
else:
    df_asist["Fecha_base"] = pd.NaT

# Filtrar Área
df_inasist = maybe_filter_area(df_inasist, only_area)
df_asist = maybe_filter_area(df_asist, only_area)

# =========================
# Turnos planificados (Activos + Turnos) para base cumplimiento
# =========================
# Asegurar columna RUT
if "RUT" not in df_activos.columns:
    rut_col_act = find_col(df_activos, ["RUT", "Rut", "rut"])
    if rut_col_act and rut_col_act != "RUT":
        df_activos = df_activos.rename(columns={rut_col_act: "RUT"})

fixed_cols = [c for c in ["Nombre del Colaborador", "RUT", "Área", "Supervisor"] if c in df_activos.columns]
date_cols = [c for c in df_activos.columns if c not in fixed_cols]

df_act_long = df_activos.melt(
    id_vars=[c for c in fixed_cols if c in df_activos.columns],
    value_vars=date_cols,
    var_name="Fecha",
    value_name="Turno_planificado"
)
df_act_long["Fecha_dt"] = df_act_long["Fecha"].apply(try_parse_date_any)
df_act_long["RUT_norm"] = df_act_long["RUT"].apply(normalize_rut) if "RUT" in df_act_long.columns else ""
df_act_long["Turno_planificado"] = df_act_long["Turno_planificado"].astype(str).str.strip()
df_act_long.loc[df_act_long["Turno_planificado"].isin(["", "nan", "NaT", "None", "-"]), "Turno_planificado"] = ""

# =========================
# Construcción Incidencias (con filtros)
# Campos requeridos:
# Fecha, Nombre, Primer Apellido, Segundo Apellido, RUT, Turno, Especialidad,
# Supervisor, Tipo_Incidencia, Detalle, Clasificación Manual
# =========================
inc_rows = []

# 1) Asistencias: SOLO si hay incidencia (Retraso o Salida) y cumple mínimo de horas (sum)
retr = get_num(df_asist, ["Retraso (horas)", "Retraso horas", "Retraso"])
sal = get_num(df_asist, ["Salida Anticipada (horas)", "Salida Anticipada", "Salida anticipada (horas)"])
suma = retr + sal

mask_asist = ((retr > 0) | (sal > 0)) & (suma >= float(min_horas))

df_asist_inc = df_asist[mask_asist].copy()

df_asist_inc["Fecha"] = df_asist_inc["Fecha_base"].dt.date
df_asist_inc["Nombre"] = safe_text_series(df_asist_inc, ["Nombre"], "")
df_asist_inc["Primer Apellido"] = safe_text_series(df_asist_inc, ["Primer Apellido", "Primer apellido"], "")
df_asist_inc["Segundo Apellido"] = safe_text_series(df_asist_inc, ["Segundo Apellido", "Segundo apellido"], "")
df_asist_inc["RUT"] = df_asist_inc[rut_col_as].astype(str)
df_asist_inc["Turno"] = safe_text_series(df_asist_inc, ["Turno"], "")
df_asist_inc["Especialidad"] = safe_text_series(df_asist_inc, ["Especialidad"], "")
df_asist_inc["Supervisor"] = safe_text_series(df_asist_inc, ["Supervisor"], "")

df_asist_inc["Tipo_Incidencia"] = "Marcaje/Turno"
df_asist_inc["Detalle"] = (
    "Retraso_h=" + retr[mask_asist].astype(str).values
    + " | SalidaAnt_h=" + sal[mask_asist].astype(str).values
    + " | Total_h=" + suma[mask_asist].astype(str).values
)

df_asist_inc["Clasificación Manual"] = "Seleccionar"

inc_rows.append(df_asist_inc[[
    "Fecha", "Nombre", "Primer Apellido", "Segundo Apellido", "RUT",
    "Turno", "Especialidad", "Supervisor",
    "Tipo_Incidencia", "Detalle", "Clasificación Manual"
]])

# 2) Inasistencias: se listan para clasificar manualmente (no aplica umbral de horas)
df_inasist_inc = df_inasist.copy()
df_inasist_inc["Fecha"] = df_inasist_inc["Fecha_base"].dt.date
df_inasist_inc["Nombre"] = safe_text_series(df_inasist_inc, ["Nombre"], "")
df_inasist_inc["Primer Apellido"] = safe_text_series(df_inasist_inc, ["Primer Apellido", "Primer apellido"], "")
df_inasist_inc["Segundo Apellido"] = safe_text_series(df_inasist_inc, ["Segundo Apellido", "Segundo apellido"], "")
df_inasist_inc["RUT"] = df_inasist_inc[rut_col_inas].astype(str)
df_inasist_inc["Turno"] = safe_text_series(df_inasist_inc, ["Turno"], "")
df_inasist_inc["Especialidad"] = safe_text_series(df_inasist_inc, ["Especialidad"], "")
df_inasist_inc["Supervisor"] = safe_text_series(df_inasist_inc, ["Supervisor"], "")

mot = safe_text_series(df_inasist_inc, ["Motivo"], "")
df_inasist_inc["Tipo_Incidencia"] = "Inasistencia"
df_inasist_inc["Detalle"] = "Motivo=" + mot
df_inasist_inc["Clasificación Manual"] = "Seleccionar"

inc_rows.append(df_inasist_inc[[
    "Fecha", "Nombre", "Primer Apellido", "Segundo Apellido", "RUT",
    "Turno", "Especialidad", "Supervisor",
    "Tipo_Incidencia", "Detalle", "Clasificación Manual"
]])

df_incidencias = pd.concat(inc_rows, ignore_index=True)

# Fechas a datetime para filtros + orden
df_incidencias["Fecha"] = pd.to_datetime(df_incidencias["Fecha"], errors="coerce")

# =========================
# Selector de fechas (rango)
# =========================
min_fecha = df_incidencias["Fecha"].min()
max_fecha = df_incidencias["Fecha"].max()

with st.sidebar:
    st.divider()
    st.subheader("Rango de fechas")
    if pd.isna(min_fecha) or pd.isna(max_fecha):
        date_range = None
        st.caption("No se detectaron fechas válidas.")
    else:
        default_start = min_fecha.date()
        default_end = max_fecha.date()
        date_range = st.date_input(
            "Desde / Hasta",
            value=(default_start, default_end),
            min_value=default_start,
            max_value=default_end
        )

if date_range and isinstance(date_range, tuple) and len(date_range) == 2:
    d1, d2 = date_range
    mask_dates = (df_incidencias["Fecha"].dt.date >= d1) & (df_incidencias["Fecha"].dt.date <= d2)
    df_incidencias = df_incidencias[mask_dates].copy()

df_incidencias = df_incidencias.sort_values(["Fecha", "RUT"], na_position="last").reset_index(drop=True)

# =========================
# UI: Editor + resumen (dinámico en la app)
# =========================
st.subheader("Detalle (Incidencias por comprobar)")

edited = st.data_editor(
    df_incidencias,
    use_container_width=True,
    num_rows="dynamic",
    column_config={
        "Fecha": st.column_config.DateColumn(format="DD-MM-YYYY"),
        "Clasificación Manual": st.column_config.SelectboxColumn(
            options=CLASIF_OPTS,
            required=True
        ),
    },
    key="editor_incidencias"
)

st.subheader("Resumen dinámico (solo Injustificada)")
df_inj = edited[edited["Clasificación Manual"] == "Injustificada"].copy()

if df_inj.empty:
    st.info("Aún no hay registros clasificados como 'Injustificada'.")
    resumen_app = pd.DataFrame(columns=["Tipo_Incidencia", "Cantidad"])
else:
    resumen_app = (
        df_inj.groupby(["Tipo_Incidencia"], dropna=False)
              .size()
              .reset_index(name="Cantidad")
              .sort_values("Cantidad", ascending=False)
    )
    st.dataframe(resumen_app, use_container_width=True)

# =========================
# Base Cumplimiento (para Excel dinámico)
# - Turnos planificados por RUT_norm (no vacíos)
# - Traer nombre/apellidos desde incidencias (si existe)
# =========================
df_act_long_valid = df_act_long[df_act_long["Turno_planificado"] != ""].copy()
turnos_plan = (
    df_act_long_valid.groupby("RUT_norm")
    .size()
    .reset_index(name="Turnos_planificados")
)

# Nombre/apellidos desde lo editado (para poblar Cumplimiento)
tmp = edited.copy()
tmp["RUT_norm"] = tmp["RUT"].apply(normalize_rut)
name_map = (
    tmp.dropna(subset=["RUT_norm"])
       .drop_duplicates("RUT_norm")[["RUT_norm", "Nombre", "Primer Apellido", "Segundo Apellido"]]
)

turnos_base = turnos_plan.merge(name_map, on="RUT_norm", how="left")
turnos_base["Nombre"] = turnos_base["Nombre"].fillna("")
turnos_base["Primer Apellido"] = turnos_base["Primer Apellido"].fillna("")
turnos_base["Segundo Apellido"] = turnos_base["Segundo Apellido"].fillna("")
turnos_base = turnos_base.sort_values(["Nombre", "Primer Apellido", "Segundo Apellido", "RUT_norm"])

# =========================
# Export Excel:
# - Agregamos columna oculta RUT_norm_hidden solo para que Cumplimiento tenga conteo exacto.
# =========================
st.subheader("Descarga (Excel dinámico)")

export_df = edited.copy()
export_df["Fecha"] = pd.to_datetime(export_df["Fecha"], errors="coerce")

# Columna técnica para fórmulas (se oculta en Excel)
export_df["RUT_norm_hidden"] = export_df["RUT"].apply(normalize_rut)

excel_bytes = to_excel_bytes_dynamic(
    incidencias_df=export_df,
    turnos_base_df=turnos_base[["Nombre", "Primer Apellido", "Segundo Apellido", "RUT_norm", "Turnos_planificados"]]
)

st.download_button(
    "Descargar Excel consolidado (dropdown + fórmulas + Cabify)",
    data=excel_bytes,
    file_name="reporte_incidencias_consolidado.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

st.caption(
    "En el Excel descargado: edita 'Clasificación Manual' (dropdown). "
    "Las hojas 'Resumen' y 'Cumplimiento' se actualizan por fórmulas."
)
